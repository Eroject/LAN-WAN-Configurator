import openpyxl as xl
import os
import re
import tkinter as tk
from tkinter import messagebox
import openpyxl



###################################################################################################

base_path = r'.\reseau_router\result'

###################################################################################################

nom_fichier_adressage = os.path.join(base_path, 'Table_adressage.xlsx')
nom_fichier_excel = os.path.join(base_path, 'Table_des_sous_reseaux.xlsx')
nom_fichier_sous_reseaux = nom_fichier_excel
nom_fichier_routage = os.path.join(base_path, 'Table_de_routage.xlsx')
interface_path = os.path.join(base_path, 'int.txt')
routage_path = os.path.join(base_path, 'rou.txt')
dhcp_path = os.path.join(base_path, 'dhc.txt')
file_path_adressage = nom_fichier_adressage
file_path_routage = nom_fichier_routage
file_path_sousReseaux = nom_fichier_excel


##########################################################################################
##fonctions

def somme_binaire(binaire1, binaire2):
    maxlen = max(len(binaire1), len(binaire2))
    binaire1 = binaire1.zfill(maxlen)
    binaire2 = binaire2.zfill(maxlen)
    resultat = ""
    retenue = 0
    for i in range(maxlen - 1, -1, -1):
        bit1 = int(binaire1[i])
        bit2 = int(binaire2[i])
        somme_bits = bit1 + bit2 + retenue
        resultat = str(somme_bits % 2) + resultat  
        retenue = somme_bits // 2  
    if retenue:
        resultat = "1" + resultat
    return resultat
def binaire_vers_decimal(nombre_binaire):
    nombre_decimal = 0
    puissance = 0
    for chiffre in reversed(str(nombre_binaire)):
        if chiffre == '1':
            nombre_decimal += 2 ** puissance
        puissance += 1
    return nombre_decimal

def decimal_vers_binaire(nombre_decimal):
    if nombre_decimal == 0:
        return "00000000"  # Cas spécial pour le nombre décimal 0
    
    nombre_binaire = ""
    nombre_binaire1 = ""
    while nombre_decimal > 0 :
        reste = nombre_decimal % 2
        nombre_binaire = str(reste) + nombre_binaire
        nombre_decimal //= 2
    n=len(nombre_binaire)
    if n<8:
        for i in range (n,8):
            nombre_binaire1=nombre_binaire1+str(0)
    return  nombre_binaire1+nombre_binaire

def prefixetomask(prefixe):
    strg=""
    binary_mask=[]
    maskL=[]
    mask=""
    for i in range (prefixe+1):
        binary_mask.append(1)
    for i in range (32-prefixe+1,32):
        binary_mask.append(0)
    for i in range (1,33):
        strg=strg+str(binary_mask[i])
        if i%8 ==0:
            strg=strg+"."
    maskL=strg.split(".")[:-1]
    for i in range (4):
        mask =mask+str(binaire_vers_decimal(int(maskL[i])))
        if i!=3:
            mask=mask+"."
    return (mask)


def network_adress(prefixe,ip_adress):
    binary_ip_adress=""
    binary_ip_adress0=""
    strg2=""
    ip_adressf=""
    L=ip_adress.split(".")

    coprefixe=32-prefixe-1
    for i in range(4):
        binary_ip_adress=binary_ip_adress+str(decimal_vers_binaire(int(L[i])))
        if i!=3:
                binary_ip_adress=binary_ip_adress

    for i in range (prefixe,32):
        binary_ip_adress0=binary_ip_adress0+str(0)
    binary_ip_adress=binary_ip_adress[:prefixe]+binary_ip_adress0
    for i in range (1,33):
        strg2=strg2+binary_ip_adress[i-1]
        if i%8 ==0:
                strg2=strg2+"."
    binary_ip_adressL=strg2.split(".")[:-1]
    for i in range (4):
        ip_adressf =ip_adressf+str(binaire_vers_decimal(int(binary_ip_adressL[i])))
        if i!=3:
            ip_adressf=ip_adressf+"."
    return(ip_adressf)
def diffusion_adress(prefixe,ip_adress):
    binary_ip_adress=""
    binary_ip_adress0=""
    strg2=""
    ip_adressf=""
    L=ip_adress.split(".")

    coprefixe=32-prefixe-1
    for i in range(4):
        binary_ip_adress=binary_ip_adress+str(decimal_vers_binaire(int(L[i])))
        if i!=3:
                binary_ip_adress=binary_ip_adress

    for i in range (prefixe,32):
        binary_ip_adress0=binary_ip_adress0+str(1)
    binary_ip_adress=binary_ip_adress[:prefixe]+binary_ip_adress0
    for i in range (1,33):
        strg2=strg2+binary_ip_adress[i-1]
        if i%8 ==0:
                strg2=strg2+"."
    binary_ip_adressL=strg2.split(".")[:-1]
    for i in range (4):
        ip_adressf =ip_adressf+str(binaire_vers_decimal(int(binary_ip_adressL[i])))
        if i!=3:
            ip_adressf=ip_adressf+"."
    return(ip_adressf)

def network_next_adress(diffusion_previous_adrs):
    diffusion_adrs=diffusion_previous_adrs
    diffusion_adrsBin_str =""
    ip_next_adress=""
    strg3=""
    diffusion_adrsBin=diffusion_adrs.split(".")
    for i in range(4):
            diffusion_adrsBin[i]=str(decimal_vers_binaire(int(diffusion_adrsBin[i])))
            diffusion_adrsBin_str=diffusion_adrsBin_str+diffusion_adrsBin[i]
    result1=somme_binaire(diffusion_adrsBin_str,'1')
    for i in range (1,33):
            strg3=strg3+result1[i-1]
            if i%8 ==0:
                    strg3=strg3+"."
    binary_ip_next_adressL=strg3.split(".")[:-1]
    for i in range (4):
        ip_next_adress =ip_next_adress+str(binaire_vers_decimal(int(binary_ip_next_adressL[i])))
        if i!=3:
            ip_next_adress=ip_next_adress+"."
    return(ip_next_adress)
###############################################


def prefixe(nbr_machine):
    n=0
    i=1
    while (n<nbr_machine):
        i=i+1
        n=2**i-2
    prefixe=32-i
    return(prefixe)

def nbr_machine_from_str(ch):
    strg=""
    cas=False
    for i in range (len(ch)):
        if ch[i]=='(':
            cas=True
        if ch[i]==')':
            cas=True 
        if cas == True:
            strg=strg+ch[i]
    return (int(strg[1:-1]))

def nom_reseau_from_str(ch):
    strg=""
    cas=False
    for i in range (len(ch)):
        if ch[i]=='(':
            cas=True
        if ch[i]==')':
            cas=True 
        if cas == True:
            strg=strg+ch[i]
    return ((strg[1:-1]))

def prefixetomask_ligne(prefixeL,ligne):
    return(prefixetomask(prefixeL[ligne-1]))

def adress_sous_reseau_ligne(ip_adress,prefixeL,ligne):#ligne commence par 1
    if ligne == 1 :
        networkadrs=ip_adress
    for i in range (ligne-1):
        prefixe=prefixeL[i]
        diffusion_previous_adrs=diffusion_adress(prefixe,ip_adress)
        networkadrs=network_next_adress(diffusion_previous_adrs)
        ip_adress=networkadrs
    return(networkadrs)

def diffusion_adress_ligne(prefixeL,ip_adress,ligne):
    return(diffusion_adress(prefixeL[ligne-1],adress_sous_reseau_ligne(ip_adress,prefixeL,ligne)))

def first_host_adress(ip_adress):
    diffusion_adrs=ip_adress
    diffusion_adrsBin_str =""
    ip_next_adress=""
    strg3=""
    diffusion_adrsBin=diffusion_adrs.split(".")
    for i in range(4):
            diffusion_adrsBin[i]=str(decimal_vers_binaire(int(diffusion_adrsBin[i])))
            diffusion_adrsBin_str=diffusion_adrsBin_str+diffusion_adrsBin[i]
            
    result1=somme_binaire(diffusion_adrsBin_str,'1')
    for i in range (1,33):
            strg3=strg3+result1[i-1]
            if i%8 ==0:
                    strg3=strg3+"."
    binary_ip_next_adressL=strg3.split(".")[:-1]
    for i in range (4):
        ip_next_adress =ip_next_adress+str(binaire_vers_decimal(int(binary_ip_next_adressL[i])))
        if i!=3:
            ip_next_adress=ip_next_adress+"."
    return(ip_next_adress)

def last_host_adress(diffusion_ligne_adress):
    diffusion_adrs=diffusion_ligne_adress
    diffusion_adrsBin_str =""
    ip_next_adress=""
    strg3=""
    diffusion_adrsBin=diffusion_adrs.split(".")
    for i in range(4):
            diffusion_adrsBin[i]=str(decimal_vers_binaire(int(diffusion_adrsBin[i])))
            diffusion_adrsBin_str=diffusion_adrsBin_str+diffusion_adrsBin[i]

    diffusion_adrsBin_str=diffusion_adrsBin_str[:-1]+"0"  
    result1=diffusion_adrsBin_str
    for i in range (1,33):
            strg3=strg3+result1[i-1]
            if i%8 ==0:
                    strg3=strg3+"."
    binary_ip_next_adressL=strg3.split(".")[:-1]
    for i in range (4):
        ip_next_adress =ip_next_adress+str(binaire_vers_decimal(int(binary_ip_next_adressL[i])))
        if i!=3:
            ip_next_adress=ip_next_adress+"."
    return(ip_next_adress)


############################################################################
def nombre_de_lignes(nom_fichier):
    classeur = xl.load_workbook(nom_fichier)
    feuille = classeur.active
    return( feuille.max_row-3)


def lire_ch_nbre_machine(nom_fichier):
    L=[]
    #dossierdata=r'C:\Users\lenovo\Desktop\exam\data'
    ligne=1
    ligne=ligne+2
    classeur = xl.load_workbook(nom_fichier)
    feuille = classeur.active
    for ligne in feuille.iter_rows(min_row=ligne,max_col=1):
        for cellule in ligne:
            L.append(cellule.value)
    classeur.close()
    return(L)

def lire_ch_nomdereseau(nom_fichier):
    L=[]
    #dossierdata=r'C:\Users\lenovo\Desktop\exam\data'
    ligne=1
    ligne=ligne+1
    classeur = xl.load_workbook(nom_fichier)
    feuille = classeur.active
    for ligne in feuille.iter_rows(min_row=ligne,min_col=2,max_col=2):
        for cellule in ligne:
            L.append(cellule.value)
    classeur.close()
    return(L)

def lire_ch_nomdereseau_tablereseau(nom_fichier):
    L=[]
    #dossierdata=r'C:\Users\lenovo\Desktop\exam\data'
    ligne=1
    ligne=ligne+2
    classeur = xl.load_workbook(nom_fichier)
    feuille = classeur.active
    for ligne in feuille.iter_rows(min_row=ligne,min_col=1,max_col=1):
        for cellule in ligne:
            L.append(cellule.value)
    classeur.close()
    return(L)

def lire_ip_adress(nom_fichier):
    L=[]
    #dossierdata=r'C:\Users\lenovo\Desktop\exam\data'
    ligne=1
    ligne=ligne+2
    classeur = xl.load_workbook(nom_fichier)
    feuille = classeur.active
    for ligne in feuille.iter_rows(min_row=ligne,max_row=ligne ,min_col=1,max_col=6):
        for cellule in ligne:
            L.append(cellule.value)
    classeur.close()
    return(L[-1])

def lire_value(nom_fichier,ligne,col):
    L=[]
    #dossierdata=r'C:\Users\lenovo\Desktop\exam\data'
    
    ligne=ligne+2
    classeur = xl.load_workbook(nom_fichier)
    feuille = classeur.active
    for ligne in feuille.iter_rows(min_row=ligne,max_row=ligne ,min_col=col,max_col=col):
        for cellule in ligne:
            L.append(cellule.value)
    classeur.close()
    return(L[-1])

def ecrire_dans_feuille_existante(nom_fichier, ligne, colonne, valeur):
    ligne=ligne+2
    #dossierdata=r'C:\Users\lenovo\Desktop\exam\data'
    #dossierResultat=r'C:\Users\lenovo\Desktop\exam\result'
    #nom_fichier_excel = os.path.join(dossierdata,nom_fichier)
    nom_fichier_excel = nom_fichier
    classeur = xl.load_workbook(nom_fichier_excel)
    feuille = classeur.active
    feuille.cell(row=ligne, column=colonne, value=valeur)
    fichiersave= os.path.join(nom_fichier)
    #classeur.save(os.path.join(dossierResultat,nom_fichier))
    classeur.save(fichiersave)
    #classeur.close()
    

def excel_to_dict(input_file):
    workbook = xl.load_workbook(input_file)
    sheet = workbook.active
    data_dict = {}
    keysL=[]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        key = row[0]  # Utilisation de la première colonne comme clé
        values = row[1:4]  # Les autres colonnes comme valeurs dans une listee
        if key in data_dict:
            data_dict[key].append(values)
        else:
            data_dict[key] = [values]
    return data_dict
def remove_content_inside_parentheses(input_string):
    result_string = re.sub(r'\([^)]*\)', '', input_string)
    return result_string
        
    
    
###################################################################################################
## Partie graphique
entries_machines = []
routeur_config = {}
ports_disponibles = ["Se 0/0", 
                     "Se 0/1", 
                     "Se 1/0", 
                     "Se 1/1", 
                     'Fa 1/0',
                     'Fa 1/1',
                     'Fa 2/0',
                     'Fa 2/1',
                     'Fa 3/0',
                     'Fa 3/1',
                     'Fa 4/0',
                     'Fa 4/1']

reseaux_existants = []
nbr_machine=[]

# Fonction pour valider le nombre de LANs et WANs
def validate_lan_wan():
    try:
        global num_lans, num_wans
        num_lans = int(entry_lan.get())
        num_wans = int(entry_wan.get())
        
        # Générer les réseaux disponibles à partir des LANs et WANs
        for i in range(1, num_lans + 1):
            reseaux_existants.append(f"LAN {i}")
        for i in range(1, num_wans + 1):
            reseaux_existants.append(f"WAN {i}")
        
        # Masquer les widgets de la première étape
        frame_lan_wan.pack_forget()
        
        # Créer les entrées pour le nombre de machines dans chaque LAN et WAN
        create_machine_entries(num_lans, num_wans)
        
    except ValueError:
        messagebox.showerror("Erreur", "Veuillez entrer des nombres valides pour les LANs et WANs")

# Fonction pour créer des champs d'entrée pour chaque LAN et WAN
def create_machine_entries(num_lans, num_wans):
    global entries_machines
    entries_machines = []
    
    for i in range(num_lans):
        frame_machine = tk.Frame(frame_content, bg="#f0f0f0")
        frame_machine.pack(fill="x", pady=5)
        frame_machine.pack(anchor="center", expand=True)

        
        label = tk.Label(frame_machine, text=f"Nombre de machines pour LAN {i+1} :", font=("Helvetica", 12), bg="#f0f0f0")
        label.pack(side="left", padx=5)
        
        entry = tk.Entry(frame_machine)
        entry.pack(side="left", fill="x", expand=True)
        entries_machines.append(entry)
    
    for i in range(num_wans):
        frame_machine = tk.Frame(frame_content, bg="#f0f0f0")
        frame_machine.pack(fill="x", pady=5)
        frame_machine.pack(anchor="center", expand=True)

        
        
        label = tk.Label(frame_machine, text=f"Nombre de machines pour WAN {i+1} :", font=("Helvetica", 12), bg="#f0f0f0")
        label.pack(side="left", padx=5)
        
        entry = tk.Entry(frame_machine)
        entry.pack(side="left", fill="x", expand=True)
        entries_machines.append(entry)
    
    # Ajouter le bouton pour valider les machines
    button_validate_machines.pack(pady=10, fill="x")

# Fonction pour valider le nombre de machines et passer à l'étape des routeurs
def validate_machines():
    try:
        machines = [int(entry.get()) for entry in entries_machines]
        
        # Masquer tous les widgets
        frame_content.pack_forget()
        
        # Demander le nombre de routeurs
        ask_routers()
        
    except ValueError:
        messagebox.showerror("Erreur", "Veuillez entrer des nombres valides pour les machines")

# Fonction pour demander le nombre de routeurs
def ask_routers():
    global entry_routers
    frame_routers.pack(fill="both", expand=True)
    frame_routers.pack(anchor="center", expand=True)

    
    label_router = tk.Label(frame_routers, text="Nombre de routeurs :", font=("Helvetica", 14))
    label_router.pack(pady=10)
    
    entry_routers = tk.Entry(frame_routers)
    entry_routers.pack(pady=5, fill="x")
    
    button_validate_routers.pack(pady=10, fill="x")

# Fonction pour valider les routeurs et passer à la configuration des ports
def validate_routers():
    try:
        num_routers = int(entry_routers.get())
        
        # Masquer l'entrée et le bouton
        frame_routers.pack_forget()
        
        # Créer la configuration des routeurs
        create_router_config(num_routers)
        
    except ValueError:
        messagebox.showerror("Erreur", "Veuillez entrer un nombre valide de routeurs")

# Fonction pour créer les Listbox pour chaque routeur
def create_router_config(num_routers):
    global listbox_ports, listbox_reseaux, entry_network, current_router, label_router
    
    current_router = 0  # Initialisation de l'index du routeur
    
    routeur_config["Routeur"] = {}
    
    frame_config.pack(anchor="center",fill="both", expand=True)   
    
    label_router = tk.Label(frame_config, text=f"Configuration du Routeur {current_router + 1}", font=("Helvetica", 14))
    label_router.pack(pady=10)
    
    frame_ports = tk.Frame(frame_config, bg="#f0f0f0")
    frame_ports.pack(fill="x", pady=5)
    frame_ports.pack(anchor="center", expand=True)

    
    label_ports = tk.Label(frame_ports, text="Sélectionnez un port :", font=("Helvetica", 12), bg="#f0f0f0")
    label_ports.pack(pady=5, side="left")
    
    # Listbox des ports
    listbox_ports = tk.Listbox(frame_ports, height=5, exportselection=False)
    for port in ports_disponibles:
        listbox_ports.insert(tk.END, port)
    listbox_ports.pack(side="left", pady=5, fill="x", expand=True)
    
    frame_reseaux = tk.Frame(frame_config, bg="#f0f0f0")
    frame_reseaux.pack(fill="x", pady=5)
    frame_reseaux.pack(anchor="center", expand=True)

    
    label_reseaux = tk.Label(frame_reseaux, text="Sélectionnez un réseau :", font=("Helvetica", 12), bg="#f0f0f0")
    label_reseaux.pack(pady=5, side="left")
    
    # Listbox des réseaux existants (LANs et WANs)
    listbox_reseaux = tk.Listbox(frame_reseaux, height=5, exportselection=False)
    for reseau in reseaux_existants:
        listbox_reseaux.insert(tk.END, reseau)
    listbox_reseaux.pack(side="left", pady=5, fill="x", expand=True)
    
    button_add_network.pack(pady=10, fill="x")
    button_next_router.pack(pady=10, fill="x")
    button_validate_all.pack(pady=10, fill="x")

# Fonction pour ajouter le réseau sélectionné à un port
def add_network_to_port():
    selected_port = listbox_ports.get(tk.ACTIVE)
    selected_reseau = listbox_reseaux.get(tk.ACTIVE)
    
    if selected_port and selected_reseau:
        if f"Routeur {current_router + 1}" not in routeur_config["Routeur"]:
            routeur_config["Routeur"][f"Routeur {current_router + 1}"] = {}
        routeur_config["Routeur"][f"Routeur {current_router + 1}"][selected_port] = selected_reseau
        messagebox.showinfo("Ajout", f"Réseau '{selected_reseau}' ajouté au {selected_port} du Routeur {current_router + 1}")
    else:
        messagebox.showerror("Erreur", "Veuillez sélectionner un port et un réseau")

# Fonction pour passer au routeur suivant
def next_router():
    global current_router
    
    if current_router < int(entry_routers.get()) - 1:
        current_router += 1
        listbox_ports.selection_clear(0, tk.END)
        listbox_reseaux.selection_clear(0, tk.END)
        label_router.config(text=f"Configuration du Routeur {current_router + 1}")
    else:
        messagebox.showinfo("Information", "Tous les routeurs ont été configurés")

# Fonction pour valider la configuration de tous les routeurs
def validate_all_routers():
    for widget in frame_config.winfo_children():
        widget.pack_forget()
    print(routeur_config)
    #messagebox.showinfo("Succès", f"Configuration terminée: {routeur_config}")
    final_function(routeur_config)

# Fonction finale
def final_function(config):
    messagebox.showinfo("Finalisation", f"Routeurs configurés: {config}")
    root.quit()


# Création de la fenêtre principale
root = tk.Tk()
root.title("Configuration des LANs, WANs et Routeurs")
root.geometry("800x600")
# Taille de la fenêtre
window_width = 800
window_height = 600

# Obtenir la largeur et la hauteur de l'écran
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

# Calculer la position pour centrer la fenêtre
position_x = (screen_width // 2) - (window_width // 2)
position_y = (screen_height // 2) - (window_height // 2)

# Définir la géométrie de la fenêtre avec position centrée
root.geometry(f"{window_width}x{window_height}+{position_x}+{position_y}")
# Création d'un cadre parent pour centrer les frames
frame_parent = tk.Frame(root, bg="#f0f0f0")
frame_parent.pack(fill="both", expand=True)

# Centrer le contenu dans le cadre parent
frame_lan_wan = tk.Frame(frame_parent, bg="#f0f0f0")
frame_lan_wan.pack(anchor="center", expand=True)

frame_content = tk.Frame(frame_parent, bg="#f0f0f0")
frame_content.pack(anchor="center", expand=True)


frame_routers = tk.Frame(frame_parent, bg="#f0f0f0")
frame_routers.pack(anchor="center", expand=True)


frame_config = tk.Frame(frame_parent, bg="#f0f0f0")
frame_config.pack(anchor="center", expand=True)


# Créer les éléments de la première étape (LANs et WANs)
label_ipAdress = tk.Label(frame_lan_wan, text="Adresse ip :", font=("Helvetica", 14))
label_ipAdress.pack(pady=10)

entry_ipAdress = tk.Entry(frame_lan_wan)
entry_ipAdress.pack(pady=5, fill="x")

label_lan = tk.Label(frame_lan_wan, text="Nombre de LANs :", font=("Helvetica", 14))
label_lan.pack(pady=10)

entry_lan = tk.Entry(frame_lan_wan)
entry_lan.pack(pady=5, fill="x")

label_wan = tk.Label(frame_lan_wan, text="Nombre de WANs :", font=("Helvetica", 14))
label_wan.pack(pady=10)

entry_wan = tk.Entry(frame_lan_wan)
entry_wan.pack(pady=5, fill="x")

button_validate_lan_wan = tk.Button(frame_lan_wan, text="Valider", command=validate_lan_wan, font=("Helvetica", 12), bg="#4CAF50", fg="white")
button_validate_lan_wan.pack(pady=20, fill="x")

# Créer les éléments pour valider les machines
button_validate_machines = tk.Button(frame_content, text="Valider Machines", command=validate_machines, font=("Helvetica", 12), bg="#4CAF50", fg="white")

# Créer les éléments pour valider les routeurs
button_validate_routers = tk.Button(frame_routers, text="Valider Routeurs", command=validate_routers, font=("Helvetica", 12), bg="#4CAF50", fg="white")

# Créer les éléments pour ajouter le réseau sélectionné à un port
button_add_network = tk.Button(frame_config, text="Ajouter Réseau", command=add_network_to_port, font=("Helvetica", 12), bg="#2196F3", fg="white")

# Créer les éléments pour passer au routeur suivant
button_next_router = tk.Button(frame_config, text="Routeur Suivant", command=next_router, font=("Helvetica", 12), bg="#FF9800", fg="white")

# Créer les éléments pour valider tous les routeurs
button_validate_all = tk.Button(frame_config, text="Valider Tout", command=validate_all_routers, font=("Helvetica", 12), bg="#F44336", fg="white")

# Lancement de l'application

root.mainloop()

ipAdress=entry_ipAdress.get()
machines = [int(entry.get()) for entry in entries_machines]
for i in range(len(machines)):
    nbr_machine.append((reseaux_existants[i],machines[i]))

##########################################################################################################################

#remplissage des fichiers excel par les données
    
data = routeur_config
file_path=file_path_adressage
# Charger le classeur Excel existant
wb = openpyxl.load_workbook(file_path)
ws = wb.active  # Utiliser la première feuille active

# Supprimer toutes les lignes existantes
ws.delete_rows(2, ws.max_row)

# Remplir les nouvelles données dans la feuille Excel
row = 2  # Commence après la ligne des titres
for routeur, interfaces in data['Routeur'].items():
    for interface, reseau in interfaces.items():
        # Nom de routeur simplifié (R1, R2, etc.)
        routeur_name = "R" + routeur.split()[-1]
        # Ajouter les données au fichier Excel
        ws.cell(row=row, column=1, value=routeur_name)
        ws.cell(row=row, column=2, value=f"{interface} ({reseau})")
        row += 1

# Sauvegarder le fichier avec les nouvelles données
wb.save(file_path)

file_path=file_path_routage


# Charger le classeur Excel existant
wb = openpyxl.load_workbook(file_path)
ws = wb.active  # Utiliser la première feuille active

# Supprimer toutes les lignes existantes
ws.delete_rows(2, ws.max_row)

# Remplir les nouvelles données dans la feuille Excel
row = 2  # Commence après la ligne des titres
for routeur, interfaces in data['Routeur'].items():
    for interface, reseau in interfaces.items():
        # Nom de routeur simplifié (R1, R2, etc.)
        routeur_name = "R" + routeur.split()[-1]
        # Ajouter les données au fichier Excel
        ws.cell(row=row, column=1, value=routeur_name)
        ws.cell(row=row, column=2, value=f"{reseau}")
        row += 1

# Sauvegarder le fichier avec les nouvelles données
wb.save(file_path)


# Dictionnaire des données
data = nbr_machine

file_path=file_path_sousReseaux


# Charger le classeur Excel existant
wb = openpyxl.load_workbook(file_path)
ws = wb.active  # Utiliser la première feuille active

# Supprimer toutes les lignes existantes
ws.delete_rows(3, ws.max_row)

# Remplir les nouvelles données dans la feuille Excel
row = 3  # Commence après la ligne des titres
ws.cell(row=row, column=6, value=ipAdress)

for reseau, nbr_machines in data:
    reseau_name=reseau+" ("+str(nbr_machines)+")"
    # Ajouter les données au fichier Excel
    ws.cell(row=row, column=1, value=reseau_name)
    row += 1

# Sauvegarder le fichier avec les nouvelles données
wb.save(file_path)
print("Fin d'ajout des données au excels")
###################################################################################################
#remplissage du fichier du table des sous reseaux

ip_adress=lire_ip_adress(nom_fichier_excel)
chL=[]
chL=lire_ch_nbre_machine(nom_fichier_excel)
prefixeL=[]
for li in range(len(chL)):
    nbr_mach=nbr_machine_from_str(chL[li])
    prefixeL.append(prefixe(nbr_mach))
for ligne in range (1,nombre_de_lignes(nom_fichier_excel)+2):
    masque=prefixetomask_ligne(prefixeL,ligne)
    adresse=adress_sous_reseau_ligne(ip_adress,prefixeL,ligne)
    adresse_diffusion=diffusion_adress_ligne(prefixeL,ip_adress,ligne)
    premiere_adresse=first_host_adress(adresse)
    derniere_adresse=last_host_adress(adresse_diffusion)
    plage_adresse=premiere_adresse + " , " +derniere_adresse
    adr=adresse +"/"+ str(prefixeL[ligne-1])
    ecrire_dans_feuille_existante(nom_fichier_excel, ligne, 2, adr) 
    ecrire_dans_feuille_existante(nom_fichier_excel, ligne, 3, masque)
    ecrire_dans_feuille_existante(nom_fichier_excel, ligne, 4, plage_adresse)
    ecrire_dans_feuille_existante(nom_fichier_excel, ligne, 5, adresse_diffusion)
    print(adr,masque,plage_adresse,adresse_diffusion)
print("fin T sous reseau")

##################################################################################################################
#remplissage du fichier du table d'adressage
 
L_sous_reseau=lire_ch_nomdereseau_tablereseau(nom_fichier_sous_reseaux)
L_adressage=lire_ch_nomdereseau(nom_fichier_adressage)
for i in range (len(L_adressage)):
    L_adressage[i]=nom_reseau_from_str(L_adressage[i])
for i in range (1,nombre_de_lignes(nom_fichier_sous_reseaux)+2):
    occ=0
    for j in range (1,len(L_adressage)+1):
        ch_sous_reseau=L_sous_reseau[i-1]
        ch_adressage=L_adressage[j-1]
        if ch_adressage in ch_sous_reseau:
            network_adress=lire_value(nom_fichier_sous_reseaux,i,2)[:-3]
            for r in range(occ):
                network_adress=first_host_adress(network_adress)
            premiere_adresse=first_host_adress(network_adress)
            occ=occ+1
            masque_reseau=lire_value(nom_fichier_sous_reseaux,i,3)
            ecrire_dans_feuille_existante(nom_fichier_adressage, j-1, 3,premiere_adresse)
            ecrire_dans_feuille_existante(nom_fichier_adressage, j-1, 4,masque_reseau)
            print(premiere_adresse,masque_reseau)
print("fin T adressage")
    
######################################################################################################################
#remplissage du fichier du table de routage

L_sous_reseau=lire_ch_nomdereseau_tablereseau(nom_fichier_sous_reseaux)
L_adressage=lire_ch_nomdereseau(nom_fichier_routage)

for i in range (1,nombre_de_lignes(nom_fichier_sous_reseaux)+2):
    for j in range (1,len(L_adressage)+1):
        ch_sous_reseau=L_sous_reseau[i-1]
        ch_adressage=L_adressage[j-1]
        if ch_adressage in ch_sous_reseau:
            network_adress=lire_value(nom_fichier_sous_reseaux,i,2)[:-3]
            masque_reseau=lire_value(nom_fichier_sous_reseaux,i,3)
            ecrire_dans_feuille_existante(nom_fichier_routage, j-1, 3,network_adress)
            ecrire_dans_feuille_existante(nom_fichier_routage, j-1, 4,masque_reseau)
            print(network_adress,masque_reseau)
print("fin T routage")

###########################################################################################################################"
#remplissage du fichier de la configuration des interfaces

input_excel_file = nom_fichier_adressage
result_dict = excel_to_dict(input_excel_file)

with open(interface_path, 'w') as file:
    for key, values in result_dict.items():
        file.write("\n"+key+"\n")
        for i in range (0,len(values)):
            interface=remove_content_inside_parentheses(values[i][0])
            reseau=values[i][1]
            masque=values[i][2]

            if "Fa" in interface:
                ch4=""+"\n"
            if "Se" in interface:
                ch4="clock rate 64000"+"\n"
            ch1="int "+interface+"\n"
            ch2="ip address "+reseau+" "+masque+"\n"
            ch3="no shutdown"+"\n"
            ch5="exit"+"\n"
            file.write(ch1)
            file.write(ch2)
            file.write(ch3)
            file.write(ch4)
            file.write(ch5)
        print(key)
    print("fin int fil")


############################################################################################################################""
#remplissage du fichier de la configuration de routage

input_excel_file = nom_fichier_routage
result_dict = excel_to_dict(input_excel_file)

with open(routage_path, 'w') as file:
    for key, values in result_dict.items():
        file.write("\n"+key+"\n")
        file.write("router rip\nversion 2\n")
        for i in range (len(values)):
            reseau=values[i][1]
            ch1="network "+reseau+"\n"
            file.write(ch1)
        print(key)
    print("fin rout fil")

###############################################################################################################################
#remplissage du fichier de la configuration du dhcp

input_excel_file = nom_fichier_routage
result_dict = excel_to_dict(input_excel_file)

with open(dhcp_path, 'w') as file:
    for key, values in result_dict.items():
        file.write("\n"+key+"\n")
        for i in range (len(values)):
            pool=values[i][0]
            reseau=values[i][1]
            masque=values[i][2]
            gateway=first_host_adress(reseau)
            ch1="ip dhcp pool "+pool+"\n"
            ch2="network " +reseau+" "+masque+"\n"
            ch3="default-router "+gateway +"\n"
            ch4="exit\n"
            file.write(ch1)
            file.write(ch2)
            file.write(ch3)
            file.write(ch4)
            
        print(key)
    print("fin dhc fil")

messagebox.showinfo("Information", "configuration terminé")